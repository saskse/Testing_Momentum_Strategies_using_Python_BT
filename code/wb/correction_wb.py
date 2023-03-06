from ast import Return
import math 
from fileinput import close
from importlib.machinery import FrozenImporter
from importlib.resources import path
from multiprocessing.sharedctypes import Value
import os
from queue import Empty
from sqlite3 import Row
from tkinter import E
from tkinter.messagebox import YES
from tkinter.tix import COLUMN
import pandas as pd
import numpy as np

import openpyxl     #library
from statistics import variance
from statistics import stdev
from scipy.stats import gmean
import numpy as np #import mean
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
import xlwings as xw
from xlwings.utils import rgb_to_int


from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
import re

points_1_1 = 6
points_2_1 = 6
points_2_2 = 6
points_3_1 = 6
points_3_2 = 6
points_4 = 6
points_5 = 15
#IA_output = {'Vorname': [],'Nachname': [], 'Punkte': []}
#df_IA_output = pd.DataFrame(data=IA_output) 

def load_workbook_range(range_string, ws, with_header=True, with_index=False, index_name=None):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    df = pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))

    if (with_header):
        df.columns = df.iloc[0]
        df = df.iloc[1:]
        #df.columns.name = "roli"

    if (with_index and index_name is not None):
        df = df.set_index(index_name, drop=True)
        print(df.columns)
        #df.index.name = "saskia"

    #print(df)
    return df 

def nteWurzel(x, n):
	return x**(1/float(n))
    
def correction(filenames):
    from main_wb import stud_number
    from main_wb import dir
    olat_name = dir.split("_")[-1]
    # Read Student File
    wb_stud = openpyxl.load_workbook(filenames, data_only=True) #read excel file
    ws_einleitung = wb_stud["Einleitung"]
    first_name = ws_einleitung.cell(row=11, column=5).value
    last_name = ws_einleitung.cell(row=13, column=5).value

    ws_eingabe_der_daten = wb_stud["Eingabe der Daten"]
    lookback_period_month = ws_eingabe_der_daten.cell(row=11, column=12).value
    lookback_period = int(lookback_period_month.split(" ",1)[0])
    print(lookback_period)
    holding_period_month = ws_eingabe_der_daten.cell(row=13, column=12).value
    holding_period = int(holding_period_month.split(" ",1)[0])
    mittel_ranking = ws_eingabe_der_daten.cell(row=22, column=11).value
    aktien_lo = ws_eingabe_der_daten.cell(row=26, column=12).value
    aktien_ls = ws_eingabe_der_daten.cell(row=28, column=12).value
    number_of_firms = 18
    riskfree = 0
    
    #print(dir)
    #print(filenames)
    ws_grunddaten_stud = wb_stud["Grunddaten"]
    #ws_grunddaten_sol = wb_stud["Grunddaten"]
    df_grunddaten_stud = load_workbook_range("C10:U261", ws_grunddaten_stud, with_index=True, index_name="Datum ")
    #ws_grunddaten_sol = wb_stud["Grunddaten"]

    ws_berechnung_mon_renditen_stud = wb_stud["Berechnung mon. Renditen"]
    df_berechnung_mon_renditen_stud = load_workbook_range("C13:U263", ws_berechnung_mon_renditen_stud, with_index=True, index_name="Datum ")
    np.round(df_berechnung_mon_renditen_stud,decimals=0)
    print(df_berechnung_mon_renditen_stud)
    #df_berechnung_mon_renditen_stud.to_excel("saskia.xlsx")

    #ws_berechnung_mon_renditen_stud_add_one = wb_stud["Berechnung mon. Renditen"]
    #df_berechnung_mon_renditen_stud_add_one = load_workbook_range("C268:U517", ws_berechnung_mon_renditen_stud_add_one, with_index=False)

    ws_ranking_stud = wb_stud["Ranking"]
    df_ranking_stud_lb = load_workbook_range("C13:U263", ws_ranking_stud, with_index=True, index_name="Datum ")
    df_ranking_stud_rank = load_workbook_range("C267:U517", ws_ranking_stud, with_index=True, index_name="Datum ")

    ws_kauf_verkaufsignal_stud = wb_stud["Kauf- & Verkaufsignal"]
    df_kauf_verkaufsignal_lo_stud = load_workbook_range("C14:U264", ws_kauf_verkaufsignal_stud, with_index=True, index_name="Datum ")
    df_kauf_verkaufsignal_ls_stud = load_workbook_range("W14:AO264", ws_kauf_verkaufsignal_stud, with_index=True, index_name="Datum ")

    ws_monatliche_portfoliorenditen_stud = wb_stud["Monatliche Portfoliorenditen"]
    df_monatliche_portfoliorenditen_stud = load_workbook_range("C13:F263", ws_monatliche_portfoliorenditen_stud, with_index=True, index_name="Datum ")

    ws_gesamtrendite_sr_stud = wb_stud["Gesamtrendite & SR"]
    df_gesamtrendite_sr_stud = load_workbook_range("D11:F20", ws_gesamtrendite_sr_stud) 
    
    
    print(df_grunddaten_stud.head())

    # Create Solution File
    wb_sol = openpyxl.load_workbook('C:\\Users\\senns\\Documents\\Uni_Stuff\\2022\\Bachelorarbeit\\Final_Take\\IA_3_HS22.xlsx', data_only=True)
    ws_grunddaten_sol = wb_sol["Grunddaten"]
    df_grunddaten_sol = load_workbook_range("C10:U261", ws_grunddaten_sol, with_index=True, index_name="Datum ")

    ws_berechnung_mon_renditen_sol = wb_sol["Berechnung mon. Renditen"]
    df_berechnung_mon_renditen_sol = load_workbook_range("C13:U263", ws_berechnung_mon_renditen_sol, with_index=True, index_name="Datum ")

    ws_ranking_sol = wb_sol["Ranking"]
    df_ranking_sol_lb = load_workbook_range("C13:U264", ws_ranking_sol, with_index=True, index_name="Datum ")
    df_ranking_sol_rank = load_workbook_range("C267:U517", ws_ranking_sol, with_index=True, index_name="Datum ")

    ws_kauf_verkaufsignal_sol = wb_sol["Kauf- & Verkaufsignal"]
    df_kauf_verkaufsignal_lo_sol = load_workbook_range("C14:U264", ws_kauf_verkaufsignal_sol, with_index=True, index_name="Datum ")
    df_kauf_verkaufsignal_ls_sol = load_workbook_range("W14:AO264", ws_kauf_verkaufsignal_sol, with_index=True, index_name="Datum ")

    ws_monatliche_portfoliorenditen_sol = wb_sol["Monatliche Portfoliorenditen"]
    df_monatliche_portfoliorenditen_sol = load_workbook_range("C13:F263", ws_monatliche_portfoliorenditen_sol, with_index=True, index_name="Datum ") 
    ws_gesamtrendite_sr_sol = wb_sol["Gesamtrendite & SR"]
    df_gesamtrendite_sr_sol = load_workbook_range("C11:F20", ws_gesamtrendite_sr_sol, with_index=True) 
    


    #Sheet monatliche Rendite
    print(df_berechnung_mon_renditen_sol.head())
    print(df_berechnung_mon_renditen_stud.head())
    df_berechnung_mon_renditen_sol = df_grunddaten_sol.pct_change().round(4)
    df_berechnung_mon_renditen_sol = df_berechnung_mon_renditen_sol.iloc[1: , :]
    df_berechnung_mon_renditen_stud = df_berechnung_mon_renditen_stud.astype(float).round(4)
    df_delta_berechnung_mon_renditen = df_berechnung_mon_renditen_stud == df_berechnung_mon_renditen_sol
    false_count = (~df_delta_berechnung_mon_renditen).sum().sum()
    points_stud_1_1 = max(points_1_1 - points_1_1/18/250*false_count,0)

    df_berechnung_mon_renditen_add_one_sol = df_berechnung_mon_renditen_sol + 1
    df_berechnung_mon_renditen_stud_add_one = df_berechnung_mon_renditen_stud + 1

    #Sheet Ranking obere tabelle Rendite gemäss lookback period
    if mittel_ranking == "geometrische Mittel":
        df_ranking_sol_lb = nteWurzel((df_berechnung_mon_renditen_stud_add_one).rolling(window=lookback_period).apply(np.prod, raw=True),lookback_period) - 1
        df_ranking_sol_lb = df_ranking_sol_lb.round(4)
        #df_ranking_stud_lb = df_ranking_stud_lb.astype(float).round(4)
        df_delta_ranking_lb = df_ranking_sol_lb.iloc[lookback_period-1:] == df_ranking_stud_lb.iloc[lookback_period-1:].astype(float).round(4)
        false_count = (~df_delta_ranking_lb).sum().sum()
        points_stud_2_1 = max(points_2_1-(points_2_1/((18*250)-((lookback_period-1)*18)))*false_count,0)
    else: 
        df_ranking_sol_lb = df_berechnung_mon_renditen_stud.rolling(window=lookback_period).mean()
        df_ranking_sol_lb = df_ranking_sol_lb.round(4)
        #df_ranking_stud_lb = df_ranking_stud_lb.astype(float).round(4)
        df_delta_ranking_lb = df_ranking_sol_lb.iloc[lookback_period-1:] == df_ranking_stud_lb.iloc[lookback_period-1:].astype(float).round(4)
        false_count = (~df_delta_ranking_lb).sum().sum()
        points_stud_2_1 = max(points_2_1-(points_2_1/((18*250)-((lookback_period-1)*18)))*false_count,0)
    
    #Sheet Ranking untere Tabelle Ranking bilden
    df_ranking_sol_rank = df_ranking_stud_lb.rank(axis=1, ascending=False, method='dense')
    df_delta_ranking_rank = df_ranking_sol_rank.iloc[lookback_period-1:] == df_ranking_stud_rank.iloc[lookback_period-1:]
    false_count = (~df_delta_ranking_rank).sum().sum()
    points_stud_2_2 = max(points_2_2-(points_2_2/((18*250)-((lookback_period-1)*18)))*false_count,0)

    #Kauf- & Verkaufsignal lo, linke Tabelle
    df_kauf_verkaufsignal_lo_stud = df_kauf_verkaufsignal_lo_stud.fillna(False)
    if df_kauf_verkaufsignal_lo_stud.iloc[lookback_period+holding_period-2].sum() == False: #when stud made a backward test
        pass
    else: df_kauf_verkaufsignal_lo_stud = df_kauf_verkaufsignal_lo_stud.shift(periods=holding_period)
    df_berechnung_mon_renditen_add_one_sol = df_berechnung_mon_renditen_stud_add_one.iloc[lookback_period-1:]
    df_kauf_verkaufsignal_lo_sol = df_kauf_verkaufsignal_lo_sol.iloc[lookback_period-1:]
    #df_ranking_stud_rank = df_ranking_stud_rank.iloc[:-lookback_period]
    df_ranking_sol_rank_shifted = df_ranking_stud_rank.shift(periods=holding_period) #compare correct rank for calculating rolling return
    df_kauf_verkaufsignal_lo_sol = (df_berechnung_mon_renditen_add_one_sol).rolling(window=holding_period).apply(np.prod, raw=True) - 1
    df_kauf_verkaufsignal_lo_sol[(df_ranking_sol_rank_shifted > aktien_lo)] = False
    #df_kauf_verkaufsignal_lo_stud = df_kauf_verkaufsignal_lo_stud.astype(float).round(4)
    df_kauf_verkaufsignal_lo_sol = df_kauf_verkaufsignal_lo_sol.iloc[holding_period:].astype(float).round(4)
    df_kauf_verkaufsignal_lo_stud = df_kauf_verkaufsignal_lo_stud.iloc[holding_period+lookback_period-1:].astype(float).round(4)
    df_delta_kauf_verkaufsignal_lo = df_kauf_verkaufsignal_lo_sol == df_kauf_verkaufsignal_lo_stud
    false_count = (~df_delta_kauf_verkaufsignal_lo).sum().sum()
    points_stud_3_1 = max(points_3_1-(points_3_1/((18*250)-((holding_period+lookback_period-1)*18)))*false_count,0)


    #Kauf- & Verkaufsignal ls, rechte Tabelle
    df_kauf_verkaufsignal_ls_stud = df_kauf_verkaufsignal_ls_stud.fillna(False)
    if df_kauf_verkaufsignal_ls_stud.iloc[lookback_period+holding_period-2].sum() == False: #when stud made a backward test
        pass
    else: 
        df_kauf_verkaufsignal_ls_stud = df_kauf_verkaufsignal_ls_stud.shift(periods=holding_period)
        df_monatliche_portfoliorenditen_stud = df_monatliche_portfoliorenditen_stud.shift(periods=holding_period)
    df_kauf_verkaufsignal_ls_sol = df_kauf_verkaufsignal_ls_sol.iloc[lookback_period-1:]
    df_kauf_verkaufsignal_ls_sol = df_berechnung_mon_renditen_add_one_sol.rolling(window=holding_period).apply(np.prod, raw=True) - 1
    df_kauf_verkaufsignal_ls_sol[(aktien_ls < df_ranking_sol_rank_shifted) & (df_ranking_sol_rank_shifted <= (number_of_firms - aktien_ls))] = False
    df_kauf_verkaufsignal_ls_sol[df_ranking_sol_rank_shifted > (number_of_firms - aktien_ls)] = df_kauf_verkaufsignal_ls_sol * -1
    #df_kauf_verkaufsignal_ls_stud = df_kauf_verkaufsignal_ls_stud.astype(float).round(4)
    df_kauf_verkaufsignal_ls_sol = df_kauf_verkaufsignal_ls_sol.iloc[holding_period:].astype(float).round(4)
    df_kauf_verkaufsignal_ls_stud = df_kauf_verkaufsignal_ls_stud.iloc[holding_period+lookback_period-1:].astype(float).round(4)
    df_delta_kauf_verkaufsignal_ls = df_kauf_verkaufsignal_ls_sol == df_kauf_verkaufsignal_ls_stud
    false_count = (~df_delta_kauf_verkaufsignal_ls).sum().sum()
    points_stud_3_2 = max(points_3_2-(points_3_2/((18*250)-((holding_period+lookback_period-1)*18)))*false_count,0)

    #Monatliche Portfoliorenditen
    df_monatliche_portfoliorenditen_sol["Long-only"] = (nteWurzel(1+df_kauf_verkaufsignal_lo_stud.mean(axis=1),lookback_period)-1).round(4)
    df_monatliche_portfoliorenditen_sol["Long-Short"] = (nteWurzel(1+df_kauf_verkaufsignal_ls_stud.mean(axis=1),lookback_period)-1).round(4)
    df_monatliche_portfoliorenditen_sol["Buy and Hold"] = (df_berechnung_mon_renditen_stud.mean(axis=1)).round(4)
    df_monatliche_portfoliorenditen_stud = df_monatliche_portfoliorenditen_stud.astype(float)
    #df_monatliche_portfoliorenditen_stud = df_monatliche_portfoliorenditen_stud.astype(float).round(4)
    df_delta_monatliche_portfoliorenditen = df_monatliche_portfoliorenditen_sol.iloc[holding_period+lookback_period-1:] == df_monatliche_portfoliorenditen_stud.iloc[holding_period+lookback_period-1:].astype(float).round(4)
    false_count = (~df_delta_monatliche_portfoliorenditen).sum().sum()
    points_stud_4 = max(points_4-(points_4/((3*250)-((holding_period+lookback_period-1)*3)))*false_count,0)


    df_monatliche_portfoliorenditen_sol["Long-only 1+r"] = df_monatliche_portfoliorenditen_sol["Long-only"]+1
    df_monatliche_portfoliorenditen_sol["Long-Short 1+r"] = df_monatliche_portfoliorenditen_sol["Long-Short"]+1
    df_monatliche_portfoliorenditen_sol["Buy and Hold 1+r"] = df_monatliche_portfoliorenditen_sol["Buy and Hold"]+1

    df_monatliche_portfoliorenditen_sol = df_monatliche_portfoliorenditen_sol.iloc[lookback_period+holding_period-1:]

    #arith Mittel
    monatl_arithm_durchschnittsrendite_sol_lo = df_monatliche_portfoliorenditen_stud["Long-only"].iloc[lookback_period+holding_period-1:].mean()
    monatl_arithm_durchschnittsrendite_sol_ls = df_monatliche_portfoliorenditen_stud["Long-Short"].mean()
    monatl_arithm_durchschnittsrendite_sol_bah = df_monatliche_portfoliorenditen_stud["Buy and Hold"].mean()
    annualisierte_arithm_rendite_sol_lo = (1+df_gesamtrendite_sr_stud.iloc[0,0])**12-1
    annualisierte_arithm_rendite_sol_ls = (1+df_gesamtrendite_sr_stud.iloc[0,1])**12-1
    annualisierte_arithm_rendite_sol_bah = (1+df_gesamtrendite_sr_stud.iloc[0,2])**12-1
    df_gmean_lo = 1+df_monatliche_portfoliorenditen_stud["Long-only"].iloc[lookback_period+holding_period-1:]
    df_gmean_ls = 1+df_monatliche_portfoliorenditen_stud["Long-Short"].iloc[lookback_period+holding_period-1:]
    df_gmean_bah = 1+df_monatliche_portfoliorenditen_stud["Buy and Hold"].iloc[lookback_period+holding_period-1:]
    monatl_geom_durchschnittsrendite_sol_lo = nteWurzel(np.prod(df_gmean_lo),250-lookback_period-holding_period-1)-1
    monatl_geom_durchschnittsrendite_sol_ls = nteWurzel(np.prod(df_gmean_ls),250-lookback_period-holding_period-1)-1
    monatl_geom_durchschnittsrendite_sol_bah = nteWurzel(np.prod(df_gmean_bah),250-lookback_period-holding_period-1)-1
    annualisierte_geom_rendite_sol_lo = (1+df_gesamtrendite_sr_stud.iloc[2,0])**12-1
    annualisierte_geom_rendite_sol_ls = (1+df_gesamtrendite_sr_stud.iloc[2,1])**12-1
    annualisierte_geom_rendite_sol_bah = (1+df_gesamtrendite_sr_stud.iloc[2,2])**12-1
    var_monatlich_sol_lo = np.var(df_monatliche_portfoliorenditen_stud["Long-only"], ddof=1)
    var_monatlich_sol_ls = np.var(df_monatliche_portfoliorenditen_stud["Long-Short"], ddof=1)
    var_monatlich_sol_bah = np.var(df_monatliche_portfoliorenditen_stud["Buy and Hold"], ddof=1)
    var_jaehrlich_sol_lo = df_gesamtrendite_sr_stud.iloc[5,0]*12
    var_jaehrlich_sol_ls = df_gesamtrendite_sr_stud.iloc[5,1]*12
    var_jaehrlich_sol_bah = df_gesamtrendite_sr_stud.iloc[5,2]*12
    vola_monatlich_sol_lo = nteWurzel(df_gesamtrendite_sr_stud.iloc[4,0],2)
    vola_monatlich_sol_ls = nteWurzel(df_gesamtrendite_sr_stud.iloc[4,1],2)
    vola_monatlich_sol_bah = nteWurzel(df_gesamtrendite_sr_stud.iloc[4,2],2)
    vola_jaehrlich_sol_lo = nteWurzel(df_gesamtrendite_sr_stud.iloc[6,0],2)
    vola_jaehrlich_sol_ls = nteWurzel(df_gesamtrendite_sr_stud.iloc[6,1],2)
    vola_jaehrlich_sol_bah = nteWurzel(df_gesamtrendite_sr_stud.iloc[6,2],2)
    sharpe_ratio_sol_ls = (annualisierte_arithm_rendite_sol_lo - riskfree)/ vola_jaehrlich_sol_lo
    sharpe_ratio_sol_lo = (annualisierte_arithm_rendite_sol_ls - riskfree)/ vola_jaehrlich_sol_ls
    sharpe_ratio_sol_bah = (annualisierte_arithm_rendite_sol_bah - riskfree)/ vola_jaehrlich_sol_bah

    gesamtrendite_sr_sol = {'Long-only': [monatl_arithm_durchschnittsrendite_sol_lo,annualisierte_arithm_rendite_sol_lo,monatl_geom_durchschnittsrendite_sol_lo,annualisierte_geom_rendite_sol_lo,var_monatlich_sol_lo,var_jaehrlich_sol_lo,vola_monatlich_sol_lo,vola_jaehrlich_sol_lo,sharpe_ratio_sol_lo], 'Long-Short': [monatl_arithm_durchschnittsrendite_sol_ls,annualisierte_arithm_rendite_sol_ls,monatl_geom_durchschnittsrendite_sol_ls,annualisierte_geom_rendite_sol_ls,var_monatlich_sol_ls,var_jaehrlich_sol_ls,vola_monatlich_sol_ls,vola_jaehrlich_sol_ls,sharpe_ratio_sol_ls], 'Buy and Hold': [monatl_arithm_durchschnittsrendite_sol_bah,annualisierte_arithm_rendite_sol_bah,monatl_geom_durchschnittsrendite_sol_bah,annualisierte_geom_rendite_sol_bah,var_monatlich_sol_bah,var_jaehrlich_sol_bah,vola_monatlich_sol_bah,vola_jaehrlich_sol_bah,sharpe_ratio_sol_bah]}
    df_gesamtrendite_sr_sol = pd.DataFrame(data=gesamtrendite_sr_sol, index=[1,2,3,4,5,6,7,8,9]) #set index like this to compare stud and solution
    df_gesamtrendite_sr_stud = df_gesamtrendite_sr_stud.astype(float)
    df_gesamtrendite_sr_stud = df_gesamtrendite_sr_stud.round(4)
    df_gesamtrendite_sr_sol = df_gesamtrendite_sr_sol.round(4)
    df_delta_gesamtrendite_sr = df_gesamtrendite_sr_sol == df_gesamtrendite_sr_stud
    false_count = (~df_delta_gesamtrendite_sr).sum().sum()
    points_stud_5 = max(points_5 - points_5/27*false_count,0)

    #df_gesamtrendite_sr_stud(data=gesamtrendite_sr_stud, index=['monatl. arithm. Durchschnittsrendite','annualisierte Rendite (arithm)','monatl geom. Durchschnittsrendite','annualisierte Rendite (geom)','Varianz p.m.','Varianz p.a.','Volatilität p.m.','Volatilität p.a.','Sharpe Ratio (rf = 0%)'])

    #Generate IA Output
    wb_IA_output = openpyxl.load_workbook('C:\\Users\\senns\\Documents\\Uni_Stuff\\2022\\Bachelorarbeit\\Final_Take\\IA_Output_WB.xlsx', data_only=True)
    ws_IA_output = wb_IA_output["IA Output"]
    df_IA_output = load_workbook_range("A1:C500", ws_IA_output, with_index=True)

    list_max_points = [points_1_1,points_2_1,points_2_2,points_3_1,points_3_2,points_4,points_5]
    max_points = sum(list_max_points)
    list_points_stud = [points_stud_1_1,points_stud_2_1,points_stud_2_2,points_stud_3_1,points_stud_3_2,points_stud_4,points_stud_5]
    points_stud = sum(list_points_stud)
    ws_IA_output.cell(row=6+stud_number, column=3).value = first_name
    ws_IA_output.cell(row=6+stud_number, column=4).value = last_name
    ws_IA_output.cell(row=6+stud_number, column=5).value = points_stud
    ws_IA_output.cell(row=6+stud_number, column=6).value = olat_name
    if points_stud >= max_points*0.4:
        ws_IA_output.cell(row=6+stud_number, column=7).value = '1' #means bestanden
    else: ws_IA_output.cell(row=6+stud_number, column=7).value = '0' #means nicht bestanden
    ws_IA_output.cell(row=6+stud_number, column=9).value = points_stud_1_1
    ws_IA_output.cell(row=6+stud_number, column=10).value = points_stud_2_1
    ws_IA_output.cell(row=6+stud_number, column=11).value = points_stud_2_2
    ws_IA_output.cell(row=6+stud_number, column=12).value = points_stud_3_1
    ws_IA_output.cell(row=6+stud_number, column=13).value = points_stud_3_2
    ws_IA_output.cell(row=6+stud_number, column=14).value = points_stud_4
    ws_IA_output.cell(row=6+stud_number, column=15).value = points_stud_5
    wb_IA_output.save('IA_Output_WB.xlsx')

    print('Ende')
    #df_monatliche_portfoliorenditen_stud_lo = nteWurzel(1+np.nanmean(df_kauf_verkaufsignal_lo_stud, axis=1),lookback_period) - 1
    #print(type(df_monatliche_portfoliorenditen_stud_lo))
    #df_monatliche_portfoliorenditen_stud_ls = nteWurzel(1+np.nanmean(df_kauf_verkaufsignal_ls_stud, axis=1),lookback_period) - 1
    #df_monatliche_portfoliorenditen_stud_bh = np.mean(df_sol_berechnung_mon_renditen)
    #df_monatliche_portfoliorenditen_stud = pd.DataFrame(df_monatliche_portfoliorenditen_stud_lo, df_monatliche_portfoliorenditen_stud_ls, df_monatliche_portfoliorenditen_stud_bh)
    #print(df_monatliche_portfoliorenditen_stud)

    # nteWurzel(1+df_kauf_verkaufsignal_lo_stud.mean(axis=1),lookback_period)-1
    #Monatliche Portfoliorenditen, ls

   #Monatliche Portfoliorenditen, buy-and-hold
    
    #öpis vom Roli
    df1 = df_berechnung_mon_renditen_add_one_sol.append(pd.Series(), ignore_index=True)
    df1 = df1.iloc[1: , :]
    #df1 = pd.concat([pd.Series(),df_berechnung_mon_renditen_add_one_sol.loc[:]])
    df_mult = pd.DataFrame(df1.values*df_berechnung_mon_renditen_add_one_sol.values, columns=df_berechnung_mon_renditen_add_one_sol.columns, index=df_berechnung_mon_renditen_add_one_sol.index)
    
    #berechnung_mon_renditen[].fill = red


    #print(df_stud.var())

    #points = 2000 * (df_solution.count().count() / df_mistakes.count().count())
    #print(points)
#df_IA_output.to_excel('IA_Output.xlsx')