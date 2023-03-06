import os
import glob


from ast import Return
import math 
from fileinput import close
from importlib.machinery import FrozenImporter
from importlib.resources import path
from multiprocessing.sharedctypes import Value
import os
from queue import Empty
from sqlite3 import Row
from tkinter import E
from tkinter.messagebox import YES
from tkinter.tix import COLUMN
import pandas as pd
import numpy as np

import openpyxl     #library
from statistics import variance
from statistics import stdev
from scipy.stats import gmean
import numpy as np #import mean
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
import xlwings as xw
from xlwings.utils import rgb_to_int


from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
import re

import correction_wb


def load_workbook_range(range_string, ws, with_header=True, with_index=False, index_name=None):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    df = pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))

    if (with_header):
        df.columns = df.iloc[0]
        df = df.iloc[1:]
        #df.columns.name = "roli"

    if (with_index and index_name is not None):
        df = df.set_index(index_name, drop=True)
        print(df.columns)
        #df.index.name = "saskia"

    #print(df)
    return df 

path = 'C:\\Users\\senns\\Documents\\Uni_Stuff\\2022\\Bachelorarbeit\\Final_Take\\files_for_correction_wb\\IA_3_2022-12-27T19-03\\ita_IA_3_2022-12-27T19-03-08_492'
stud_number = 0
for dir in os.listdir(path): 
    filenames = glob.glob(os.path.join(path + '\\' + dir + '\\2_submissions\\' , "*.xlsx"))
    stud_number +=1
    if filenames != []:

        wb_IA_output = openpyxl.load_workbook('C:\\Users\\senns\\Documents\\Uni_Stuff\\2022\\Bachelorarbeit\\Final_Take\\IA_Output_WB.xlsx', data_only=True)
        ws_IA_output = wb_IA_output["IA Output"]
        df_IA_output = load_workbook_range("A1:C500", ws_IA_output, with_index=True)
        olat_name = dir.split("_")[-1]
        ws_IA_output.cell(row=6+stud_number, column=6).value = olat_name
        ws_IA_output.cell(row=6+stud_number, column=5).value = 'nicht abgegeben'
        ws_IA_output.cell(row=6+stud_number, column=7).value = '0' #means nicht bestanden

        correction_wb.correction(filenames[0])
        # Code