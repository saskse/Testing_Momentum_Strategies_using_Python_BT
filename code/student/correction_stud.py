#Python library imports
import math 
import os
import pandas as pd
import numpy as np

import openpyxl     #library
from statistics import variance
from statistics import stdev
from scipy.stats import gmean
import numpy as np #import mean
import os


from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
import re

#def load workbook range function
def load_workbook_range(range_string, ws, with_header=True, with_index=False, index_name=None):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    df = pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))

    if (with_header):
        df.columns = df.iloc[0]
        df = df.iloc[1:]

    if (with_index and index_name is not None):
        df = df.set_index(index_name, drop=True)
        print(df.columns)

    return df 

#def nthRoot function
def nthRoot(x, n):
	return x**(1/float(n))

#def corretion function
def correction(filenames):
    #import directory and stud_number from dir_stud.py
    from dir_stud import dir
    from dir_stud import stud_number
    
    #set path
    path = 'C:\\Users\\senns\\Documents\\Uni_Stuff\\2022\\Bachelorarbeit\\Final_Take\\Testing_Momentum_Strategies_using_Python_BT\\data'
    
    
    #Read student file
    wb_stud = openpyxl.load_workbook(filenames, data_only=True) #load excel into Python
    olat_name = dir.split("_")[-1] #save OLAT name from dir path
    ws_einleitung = wb_stud["Einleitung"] #save sheet "Einleitung"
    #Save variables from sheet "Einleitung"
    first_name = ws_einleitung.cell(row=11, column=5).value
    last_name = ws_einleitung.cell(row=13, column=5).value
    matriculation_number = ws_einleitung.cell(row=15, column=5).value

    ws_eingabe_der_daten = wb_stud["Eingabe der Daten"] #save sheet "Eingabe der Daten"
    #Save variables from sheet "Eingabe der Daten"
    lookback_period_month = ws_eingabe_der_daten.cell(row=11, column=12).value
    lookback_period = int(lookback_period_month.split(" ",1)[0])
    holding_period_month = ws_eingabe_der_daten.cell(row=13, column=12).value
    holding_period = int(holding_period_month.split(" ",1)[0])
    mittel_ranking = ws_eingabe_der_daten.cell(row=22, column=11).value
    aktien_lo = ws_eingabe_der_daten.cell(row=26, column=12).value
    aktien_ls = ws_eingabe_der_daten.cell(row=28, column=12).value
    
    #Define variables
    number_of_firms = 18
    time_period = 250

    riskfree = 0

    points_1_1 = 6
    points_2_1 = 6
    points_2_2 = 6
    points_3_1 = 6
    points_3_2 = 6
    points_4 = 6
    points_5 = 15
    
    #Save worksheets and load workbook range of student file
    ws_grunddaten_stud = wb_stud["Grunddaten"]
    df_grunddaten_stud = load_workbook_range("C10:U261", ws_grunddaten_stud, with_index=True, index_name="Datum ")

    ws_berechnung_mon_renditen_stud = wb_stud["Berechnung mon. Renditen"]
    df_berechnung_mon_renditen_stud = load_workbook_range("C13:U263", ws_berechnung_mon_renditen_stud, with_index=True, index_name="Datum ")

    ws_ranking_stud = wb_stud["Ranking"]
    df_ranking_stud_lb = load_workbook_range("C13:U263", ws_ranking_stud, with_index=True, index_name="Datum ")
    df_ranking_stud_rank = load_workbook_range("C267:U517", ws_ranking_stud, with_index=True, index_name="Datum ")

    ws_kauf_verkaufsignal_stud = wb_stud["Kauf- & Verkaufsignal"]
    df_kauf_verkaufsignal_lo_stud = load_workbook_range("C14:U264", ws_kauf_verkaufsignal_stud, with_index=True, index_name="Datum ")
    df_kauf_verkaufsignal_ls_stud = load_workbook_range("W14:AO264", ws_kauf_verkaufsignal_stud, with_index=True, index_name="Datum ")

    ws_monatliche_portfoliorenditen_stud = wb_stud["Monatliche Portfoliorenditen"]
    df_monatliche_portfoliorenditen_stud = load_workbook_range("C13:F263", ws_monatliche_portfoliorenditen_stud, with_index=True, index_name="Datum ")

    ws_gesamtrendite_sr_stud = wb_stud["Gesamtrendite & SR"]
    df_gesamtrendite_sr_stud = load_workbook_range("D11:F20", ws_gesamtrendite_sr_stud) 
    

    #Load empty solution file
    wb_sol = openpyxl.load_workbook(path + '\\input\\IA_3_HS22_shifted.xlsx', data_only=True) #load empty solution file
    #Save worksheets and load workbook range of solution file
    ws_grunddaten_sol = wb_sol["Grunddaten"]
    df_grunddaten_sol = load_workbook_range("C10:U261", ws_grunddaten_sol, with_index=True, index_name="Datum ")

    ws_berechnung_mon_renditen_sol = wb_sol["Berechnung mon. Renditen"]
    df_berechnung_mon_renditen_sol = load_workbook_range("C13:U263", ws_berechnung_mon_renditen_sol, with_index=True, index_name="Datum ")

    ws_ranking_sol = wb_sol["Ranking"]
    df_ranking_sol_lb = load_workbook_range("C13:U264", ws_ranking_sol, with_index=True, index_name="Datum ")
    df_ranking_sol_rank = load_workbook_range("C267:U517", ws_ranking_sol, with_index=True, index_name="Datum ")

    ws_kauf_verkaufsignal_sol = wb_sol["Kauf- & Verkaufsignal"]
    df_kauf_verkaufsignal_lo_sol = load_workbook_range("C14:U264", ws_kauf_verkaufsignal_sol, with_index=True, index_name="Datum ")
    df_kauf_verkaufsignal_ls_sol = load_workbook_range("W14:AO264", ws_kauf_verkaufsignal_sol, with_index=True, index_name="Datum ")

    ws_monatliche_portfoliorenditen_sol = wb_sol["Monatliche Portfoliorenditen"]
    df_monatliche_portfoliorenditen_sol = load_workbook_range("C13:F263", ws_monatliche_portfoliorenditen_sol, with_index=True, index_name="Datum ") 
    ws_gesamtrendite_sr_sol = wb_sol["Gesamtrendite & SR"]
    df_gesamtrendite_sr_sol = load_workbook_range("C11:F20", ws_gesamtrendite_sr_sol, with_index=True) 
    

    #Sheet "Monatliche Rendite"
    df_berechnung_mon_renditen_sol = df_grunddaten_sol.pct_change().round(4) #calculate monthly return and round to 4 decimal points
    df_berechnung_mon_renditen_sol = df_berechnung_mon_renditen_sol.iloc[1: , :] #get rid of first row bc of shifted "Grunddaten" from solution file
    df_berechnung_mon_renditen_stud = df_berechnung_mon_renditen_stud.astype(float).round(4) #convert values to a float and round to 4 decimal points
    df_delta_berechnung_mon_renditen = df_berechnung_mon_renditen_stud == df_berechnung_mon_renditen_sol #built delta of student and solution dataframe
    false_count = (~df_delta_berechnung_mon_renditen).sum().sum() #count values in delta dataframe that are not equal and sum it up
    points_stud_1_1 = points_1_1 - points_1_1/number_of_firms/time_period*false_count #set points for first exercise

    df_berechnung_mon_renditen_add_one_sol = df_berechnung_mon_renditen_sol + 1 #plus one for geo medium
    df_berechnung_mon_renditen_stud_add_one = df_berechnung_mon_renditen_stud + 1 #plus one for geo medium


    #Sheet "Ranking", upper table, calculate return based on look-back period
    if mittel_ranking == "geometrische Mittel": #check if student choose geometric or arithmetic mean
        df_ranking_sol_lb = nthRoot((df_berechnung_mon_renditen_stud_add_one).rolling(window=lookback_period).apply(np.prod, raw=True),lookback_period) - 1 #calculate geometric mean
    else: 
        df_ranking_sol_lb = df_berechnung_mon_renditen_stud.rolling(window=lookback_period).mean() #calculate arithmetic mean
    df_delta_ranking_lb = df_ranking_sol_lb.iloc[lookback_period-1:].round(4) == df_ranking_stud_lb.iloc[lookback_period-1:].astype(float).round(4) #get rid of empty rows using iloc
    false_count = (~df_delta_ranking_lb).sum().sum()
    points_stud_2_1 = points_2_1-(points_2_1/((number_of_firms*time_period)-((lookback_period-1)*number_of_firms)))*false_count
    

    #Sheet "Ranking", bottom table, build ranking
    df_ranking_sol_rank = df_ranking_stud_lb.rank(axis=1, ascending=False, method='dense') #build ranking
    df_delta_ranking_rank = df_ranking_sol_rank.iloc[lookback_period-1:] == df_ranking_stud_rank.iloc[lookback_period-1:]
    false_count = (~df_delta_ranking_rank).sum().sum()
    points_stud_2_2 = points_2_2-(points_2_2/((number_of_firms*time_period)-((lookback_period-1)*number_of_firms)))*false_count


    #Sheet "Kauf- & Verkaufsignal", right table, long-only portfolio
    df_kauf_verkaufsignal_lo_stud = df_kauf_verkaufsignal_lo_stud.fillna(False) #fill empty cells with false
    if df_kauf_verkaufsignal_lo_stud.iloc[lookback_period+holding_period-2].sum() == False: #when student made a backward test
        pass
    else: df_kauf_verkaufsignal_lo_stud = df_kauf_verkaufsignal_lo_stud.shift(periods=holding_period) #when student did, shift values by holding period
    df_berechnung_mon_renditen_add_one_sol = df_berechnung_mon_renditen_stud_add_one.iloc[lookback_period-1:]
    df_kauf_verkaufsignal_lo_sol = df_kauf_verkaufsignal_lo_sol.iloc[lookback_period-1:]
    df_ranking_sol_rank_shifted = df_ranking_stud_rank.shift(periods=holding_period) #compare correct rank for calculating rolling return
    df_kauf_verkaufsignal_lo_sol = (df_berechnung_mon_renditen_add_one_sol).rolling(window=holding_period).apply(np.prod, raw=True) - 1
    df_kauf_verkaufsignal_lo_sol[(df_ranking_sol_rank_shifted > aktien_lo)] = False
    df_kauf_verkaufsignal_lo_sol = df_kauf_verkaufsignal_lo_sol.iloc[holding_period:].astype(float).round(4)
    df_kauf_verkaufsignal_lo_stud = df_kauf_verkaufsignal_lo_stud.iloc[holding_period+lookback_period-1:].astype(float).round(4)
    df_delta_kauf_verkaufsignal_lo = df_kauf_verkaufsignal_lo_sol == df_kauf_verkaufsignal_lo_stud
    false_count = (~df_delta_kauf_verkaufsignal_lo).sum().sum()
    points_stud_3_1 = points_3_1-(points_3_1/((number_of_firms*time_period)-((holding_period+lookback_period-1)*number_of_firms)))*false_count


    #Sheet "Kauf- & Verkaufsignal", left table, long-short portfolio
    df_kauf_verkaufsignal_ls_stud = df_kauf_verkaufsignal_ls_stud.fillna(False)
    if df_kauf_verkaufsignal_ls_stud.iloc[lookback_period+holding_period-2].sum() == False: #when stud made a backward test
        pass
    else: 
        df_kauf_verkaufsignal_ls_stud = df_kauf_verkaufsignal_ls_stud.shift(periods=holding_period)
        df_monatliche_portfoliorenditen_stud = df_monatliche_portfoliorenditen_stud.shift(periods=holding_period)
    df_kauf_verkaufsignal_ls_sol = df_kauf_verkaufsignal_ls_sol.iloc[lookback_period-1:]
    df_kauf_verkaufsignal_ls_sol = df_berechnung_mon_renditen_add_one_sol.rolling(window=holding_period).apply(np.prod, raw=True) - 1
    df_kauf_verkaufsignal_ls_sol[(aktien_ls < df_ranking_sol_rank_shifted) & (df_ranking_sol_rank_shifted <= (number_of_firms - aktien_ls))] = False
    df_kauf_verkaufsignal_ls_sol[df_ranking_sol_rank_shifted > (number_of_firms - aktien_ls)] = df_kauf_verkaufsignal_ls_sol * -1
    df_kauf_verkaufsignal_ls_sol = df_kauf_verkaufsignal_ls_sol.iloc[holding_period:].astype(float).round(4)
    df_kauf_verkaufsignal_ls_stud = df_kauf_verkaufsignal_ls_stud.iloc[holding_period+lookback_period-1:].astype(float).round(4)
    df_delta_kauf_verkaufsignal_ls = df_kauf_verkaufsignal_ls_sol == df_kauf_verkaufsignal_ls_stud
    false_count = (~df_delta_kauf_verkaufsignal_ls).sum().sum()
    points_stud_3_2 = points_3_2-(points_3_2/((number_of_firms*time_period)-((holding_period+lookback_period-1)*number_of_firms)))*false_count


    #Sheet "Monatliche Portfoliorenditen"
    df_monatliche_portfoliorenditen_sol["Long-only"] = (nthRoot(1+df_kauf_verkaufsignal_lo_stud.mean(axis=1),holding_period)-1).round(4) #long-only strategy
    df_monatliche_portfoliorenditen_sol["Long-Short"] = (nthRoot(1+df_kauf_verkaufsignal_ls_stud.mean(axis=1),holding_period)-1).round(4) #long-short strategy
    df_monatliche_portfoliorenditen_sol["Buy and Hold"] = (df_berechnung_mon_renditen_stud.mean(axis=1)).round(4) #buy-and-hold strategy
    
    df_monatliche_portfoliorenditen_stud = df_monatliche_portfoliorenditen_stud.astype(float)
    df_delta_monatliche_portfoliorenditen = df_monatliche_portfoliorenditen_sol.iloc[holding_period+lookback_period-1:] == df_monatliche_portfoliorenditen_stud.iloc[holding_period+lookback_period-1:].astype(float).round(4)
    false_count = (~df_delta_monatliche_portfoliorenditen).sum().sum()
    points_stud_4 = points_4-(points_4/(3*time_period-(holding_period+lookback_period-1)*3))*false_count

    df_monatliche_portfoliorenditen_sol["Long-only 1+r"] = df_monatliche_portfoliorenditen_sol["Long-only"]+1
    df_monatliche_portfoliorenditen_sol["Long-Short 1+r"] = df_monatliche_portfoliorenditen_sol["Long-Short"]+1
    df_monatliche_portfoliorenditen_sol["Buy and Hold 1+r"] = df_monatliche_portfoliorenditen_sol["Buy and Hold"]+1

    df_monatliche_portfoliorenditen_sol = df_monatliche_portfoliorenditen_sol.iloc[lookback_period+holding_period-1:]


    #Sheet "Gesamtrendite und SR"
    monatl_arithm_durchschnittsrendite_sol_lo = df_monatliche_portfoliorenditen_stud["Long-only"].iloc[lookback_period+holding_period-1:].mean() #long-only
    monatl_arithm_durchschnittsrendite_sol_ls = df_monatliche_portfoliorenditen_stud["Long-Short"].iloc[lookback_period+holding_period-1:].mean() #long-short
    monatl_arithm_durchschnittsrendite_sol_bah = df_monatliche_portfoliorenditen_stud["Buy and Hold"].iloc[lookback_period+holding_period-1:].mean() #buy-and-hold
    annualisierte_arithm_rendite_sol_lo = (1+df_gesamtrendite_sr_stud.iloc[0,0])**12-1
    annualisierte_arithm_rendite_sol_ls = (1+df_gesamtrendite_sr_stud.iloc[0,1])**12-1
    annualisierte_arithm_rendite_sol_bah = (1+df_gesamtrendite_sr_stud.iloc[0,2])**12-1
    df_gmean_lo = 1+df_monatliche_portfoliorenditen_stud["Long-only"].iloc[lookback_period+holding_period-1:]
    df_gmean_ls = 1+df_monatliche_portfoliorenditen_stud["Long-Short"].iloc[lookback_period+holding_period-1:]
    df_gmean_bah = 1+df_monatliche_portfoliorenditen_stud["Buy and Hold"].iloc[lookback_period+holding_period-1:]
    monatl_geom_durchschnittsrendite_sol_lo = nthRoot(np.prod(df_gmean_lo),time_period-lookback_period-holding_period-1)-1
    monatl_geom_durchschnittsrendite_sol_ls = nthRoot(np.prod(df_gmean_ls),time_period-lookback_period-holding_period-1)-1
    monatl_geom_durchschnittsrendite_sol_bah = nthRoot(np.prod(df_gmean_bah),time_period-lookback_period-holding_period-1)-1
    annualisierte_geom_rendite_sol_lo = (1+df_gesamtrendite_sr_stud.iloc[2,0])**12-1
    annualisierte_geom_rendite_sol_ls = (1+df_gesamtrendite_sr_stud.iloc[2,1])**12-1
    annualisierte_geom_rendite_sol_bah = (1+df_gesamtrendite_sr_stud.iloc[2,2])**12-1
    var_monatlich_sol_lo = np.var(df_monatliche_portfoliorenditen_stud["Long-only"], ddof=1)
    var_monatlich_sol_ls = np.var(df_monatliche_portfoliorenditen_stud["Long-Short"], ddof=1)
    var_monatlich_sol_bah = np.var(df_monatliche_portfoliorenditen_stud["Buy and Hold"], ddof=1)
    var_jaehrlich_sol_lo = df_gesamtrendite_sr_stud.iloc[5,0]*12
    var_jaehrlich_sol_ls = df_gesamtrendite_sr_stud.iloc[5,1]*12
    var_jaehrlich_sol_bah = df_gesamtrendite_sr_stud.iloc[5,2]*12
    vola_monatlich_sol_lo = nthRoot(df_gesamtrendite_sr_stud.iloc[4,0],2)
    vola_monatlich_sol_ls = nthRoot(df_gesamtrendite_sr_stud.iloc[4,1],2)
    vola_monatlich_sol_bah = nthRoot(df_gesamtrendite_sr_stud.iloc[4,2],2)
    vola_jaehrlich_sol_lo = nthRoot(df_gesamtrendite_sr_stud.iloc[6,0],2)
    vola_jaehrlich_sol_ls = nthRoot(df_gesamtrendite_sr_stud.iloc[6,1],2)
    vola_jaehrlich_sol_bah = nthRoot(df_gesamtrendite_sr_stud.iloc[6,2],2)
    sharpe_ratio_sol_ls = (annualisierte_arithm_rendite_sol_lo - riskfree)/ vola_jaehrlich_sol_lo
    sharpe_ratio_sol_lo = (annualisierte_arithm_rendite_sol_ls - riskfree)/ vola_jaehrlich_sol_ls
    sharpe_ratio_sol_bah = (annualisierte_arithm_rendite_sol_bah - riskfree)/ vola_jaehrlich_sol_bah

    gesamtrendite_sr_sol = {'Long-only': [monatl_arithm_durchschnittsrendite_sol_lo,annualisierte_arithm_rendite_sol_lo,monatl_geom_durchschnittsrendite_sol_lo,annualisierte_geom_rendite_sol_lo,var_monatlich_sol_lo,var_jaehrlich_sol_lo,vola_monatlich_sol_lo,vola_jaehrlich_sol_lo,sharpe_ratio_sol_lo], 'Long-Short': [monatl_arithm_durchschnittsrendite_sol_ls,annualisierte_arithm_rendite_sol_ls,monatl_geom_durchschnittsrendite_sol_ls,annualisierte_geom_rendite_sol_ls,var_monatlich_sol_ls,var_jaehrlich_sol_ls,vola_monatlich_sol_ls,vola_jaehrlich_sol_ls,sharpe_ratio_sol_ls], 'Buy and Hold': [monatl_arithm_durchschnittsrendite_sol_bah,annualisierte_arithm_rendite_sol_bah,monatl_geom_durchschnittsrendite_sol_bah,annualisierte_geom_rendite_sol_bah,var_monatlich_sol_bah,var_jaehrlich_sol_bah,vola_monatlich_sol_bah,vola_jaehrlich_sol_bah,sharpe_ratio_sol_bah]}
    df_gesamtrendite_sr_sol = pd.DataFrame(data=gesamtrendite_sr_sol, index=[1,2,3,4,5,6,7,8,9]) #set index like this to compare stud and solution
    df_gesamtrendite_sr_stud = df_gesamtrendite_sr_stud.astype(float).round(4)
    df_gesamtrendite_sr_sol = df_gesamtrendite_sr_sol.round(4)
    df_delta_gesamtrendite_sr = df_gesamtrendite_sr_sol == df_gesamtrendite_sr_stud
    false_count = (~df_delta_gesamtrendite_sr).sum().sum()
    points_stud_5 = points_5 - points_5/27*false_count #bc 27 measurements


    #Generate IA Output
    wb_IA_output = openpyxl.load_workbook(path + '\\input\\IA_Output_empty_stud.xlsx', data_only=True) #load empty IA Output file
    ws_IA_output = wb_IA_output["IA Output"]

    list_max_points = [points_1_1,points_2_1,points_2_2,points_3_1,points_3_2,points_4,points_5]
    max_points = sum(list_max_points)
    list_points_stud = [points_stud_1_1,points_stud_2_1,points_stud_2_2,points_stud_3_1,points_stud_3_2,points_stud_4,points_stud_5]
    points_stud = sum(list_points_stud)
    ws_IA_output.cell(row=6+stud_number, column=3).value = matriculation_number
    ws_IA_output.cell(row=6+stud_number, column=4).value = first_name
    ws_IA_output.cell(row=6+stud_number, column=5).value = last_name
    ws_IA_output.cell(row=6+stud_number, column=6).value = points_stud
    ws_IA_output.cell(row=6+stud_number, column=7).value = olat_name
    if points_stud >= max_points*0.4: #passing limit
        ws_IA_output.cell(row=6+stud_number, column=8).value = '1' #means passed
    else: ws_IA_output.cell(row=6+stud_number, column=8).value = '0' #means failed
    ws_IA_output.cell(row=6+stud_number, column=10).value = points_stud_1_1
    ws_IA_output.cell(row=6+stud_number, column=11).value = points_stud_2_1
    ws_IA_output.cell(row=6+stud_number, column=12).value = points_stud_2_2
    ws_IA_output.cell(row=6+stud_number, column=13).value = points_stud_3_1
    ws_IA_output.cell(row=6+stud_number, column=14).value = points_stud_3_2
    ws_IA_output.cell(row=6+stud_number, column=15).value = points_stud_4
    ws_IA_output.cell(row=6+stud_number, column=16).value = points_stud_5
    wb_IA_output.save(path + '\\output\\IA_Output_stud.xlsx')
